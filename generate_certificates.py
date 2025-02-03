import sys
import openpyxl
import os
import time
import logging
from datetime import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from PyPDF2 import PdfReader, PdfWriter
import arabic_reshaper
from bidi.algorithm import get_display
import qrcode
from concurrent.futures import ThreadPoolExecutor

# Ensure a file path argument is provided
if len(sys.argv) < 2:
    print("Error: No Excel file provided")
    sys.exit(1)

# Get the uploaded Excel file path from command-line argument
excel_file = sys.argv[1]

if not os.path.exists(excel_file):
    print(f"Error: File '{excel_file}' not found")
    sys.exit(1)

# Setup logging
def setup_logging():
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file = f"certificates_{current_datetime}.log"
    logging.basicConfig(filename=log_file, level=logging.INFO)
    logging.info("Started generating PDFs.")

setup_logging()

# Start the timer
start_time = time.time()

# Register Arabic-supporting font
font_path = os.path.join("Amiri", "Amiri-Regular.ttf")
pdfmetrics.registerFont(TTFont("Amiri", font_path))

# Load Excel file
wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb.active

# Load PDF template (assuming it exists in the same folder)
pdf_template = "certificate.pdf"

# Ensure output directory exists
output_folder = "output"
os.makedirs(output_folder, exist_ok=True)

# Define coordinates for text placement
coordinates = {
    "Référence de certificat": (229.0, 650.0),
    "Date de création": (46.0, 600.0),
    "Numéro chez le déclarant": (180.0, 600.0),
    "YYYY de Date de paiement": (300.0, 600.0),
    "Date de paiement": (440.0, 600.0),
    "Declarant": (229.0, 545),
    "Id Aziza": (229.0, 533),
    "Nom et prenom ou raison sociale": (229.0, 510.0),
    "Adresse Aziza": (229.0, 495),
    "Preffession Aziza": (229.0, 475),
    "Identifiant du bénéficiaire": (229.0, 402.0),
    "Raison social du bénéficiaire": (229.0, 390.0),
    "type identifiant ben": (229.0, 420),
    "adresse benificiaire": (229.0, 360),
    "activite benificiaire": (229.0, 350),
    "totalMontantHT": (205.0, 260.0),
    "TVA_due": (260.0, 260.0),
    "Montant total TVA comprise - Total": (307.0, 245.0),
    "Montant total TVA comprise": (307.0, 260.0),
    "TVA retenue à la source": (360.0, 260.0),
    "Montant de la retenue": (461.0, 260.0),
    "Taux de la retenue": (408.0, 260.0),
    "Montant servie": (510.0, 260.0),
    "Montant servie - Total": (510.0, 245.0),
    "QR Code": (40.0, 120.0),
}

# Cache reshaped Arabic text
reshaped_cache = {}

def reshape_arabic(text):
    if text not in reshaped_cache:
        reshaped_cache[text] = get_display(arabic_reshaper.reshape(text))
    return reshaped_cache[text]

def write_text_on_pdf(c, text, coordinates, font="Helvetica", font_size=8):
    if any("\u0600" <= char <= "\u06FF" for char in text):  # Arabic check
        bidi_text = reshape_arabic(text)
        c.setFont("Amiri", font_size)
    else:
        c.setFont(font, font_size)
        bidi_text = text
    x, y = coordinates
    c.drawString(x, y, bidi_text)

# Function to generate a PDF for a single row
def generate_pdf(row, total_files, index, start_time):
    try:
        if not any(row):  # Skip empty rows
            logging.error(f"Skipped row {index + 2}: row is empty or undefined.")
            print(f"Skipping empty row {index + 2}")
            return

        row = [str(cell).strip() if isinstance(cell, str) else cell for cell in row]

        reader = PdfReader(pdf_template)
        writer = PdfWriter()
        page = reader.pages[0]

        packet = BytesIO()
        can = canvas.Canvas(packet)

        # Write text fields from the Excel row
        write_text_on_pdf(can, str(row[3]), coordinates["Référence de certificat"], font_size=8)
        write_text_on_pdf(can, str(row[22]), coordinates["Date de création"], font_size=8)
        write_text_on_pdf(can, str(row[2]), coordinates["Numéro chez le déclarant"], font_size=8)
        write_text_on_pdf(can, str(row[4].split("-")[2]), coordinates["YYYY de Date de paiement"], font_size=8)
        write_text_on_pdf(can, str(row[4]), coordinates["Date de paiement"], font_size=8)

        # Generate QR code
        qr_text = f"{row[0]}#{row[4].split('-')[2]}#{row[4].split('-')[1]}#1#{row[3]}"
        qr = qrcode.QRCode(box_size=10, border=2)
        qr.add_data(qr_text)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_byte_stream = BytesIO()
        qr_img.save(qr_byte_stream, format="PNG")
        qr_byte_stream.seek(0)
        qr_image = ImageReader(qr_byte_stream)
        qr_x, qr_y = coordinates["QR Code"]
        can.drawImage(qr_image, qr_x, qr_y, width=100, height=100)

        can.save()

        packet.seek(0)
        overlay_pdf = PdfReader(packet)
        page.merge_page(overlay_pdf.pages[0])
        writer.add_page(page)

        # Save the output PDF
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_pdf_name = os.path.join(output_folder, f"certificat_RS_{row[17]}_{current_datetime}.pdf")
        with open(output_pdf_name, "wb") as output_pdf:
            writer.write(output_pdf)

        elapsed_time = time.time() - start_time
        avg_time_per_file = elapsed_time / (index + 1)
        remaining_files = total_files - (index + 1)
        remaining_time = avg_time_per_file * remaining_files

        print(f"\rProcessing file {index + 1} of {total_files} - Estimated time remaining: {remaining_time / 60:.2f} minutes", end="")

        logging.info(f"Generated PDF for {row[3]}")

    except Exception as e:
        logging.error(f"Error processing row {index + 2}: {str(e)}")

# Process files
def process_files():
    total_files = sheet.max_row - 1
    with ThreadPoolExecutor() as executor:
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            generate_pdf(row, total_files, index, start_time)

# Start processing
process_files()
