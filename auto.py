import openpyxl
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
import os
import arabic_reshaper
from bidi.algorithm import get_display
import qrcode
import time
from datetime import datetime
import logging
import sys
import json
import websocket
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import zipfile
import shutil

# Thread-safe Progress Tracker
class ProgressTracker:
    def __init__(self, total_files):
        self.total_files = total_files
        self.current_progress = 0
        self.processed_files = 0
        self.lock = threading.Lock()

    def update_progress(self, processed_increment=1):
        with self.lock:
            self.processed_files += processed_increment
            new_progress = min(100, (self.processed_files / self.total_files) * 100)
            
            # Only send update if progress has changed significantly
            if new_progress > self.current_progress + 0.5:  # Send update every 0.5% change
                self.current_progress = new_progress
                return self.current_progress
        return None

# Global WebSocket connection
global_ws = None
ws_lock = threading.Lock()

def on_open(ws):
    print("WebSocket connection opened")

def on_message(ws, message):
    print(f"Received message: {message}")

def on_error(ws, error):
    print(f"WebSocket error: {error}")

def on_close(ws, close_status_code, close_msg):
    print("WebSocket connection closed")

# Setup WebSocket connection
def setup_websocket(socket_id):
    global global_ws
    try:
        ws = websocket.WebSocketApp(
            f'ws://localhost:3008/progress/{socket_id}',
            on_open=on_open,
            on_message=on_message,
            on_error=on_error,
            on_close=on_close
        )
        
        wst = threading.Thread(target=ws.run_forever)
        wst.daemon = True
        wst.start()
        
        with ws_lock:
            global_ws = ws
        
        return ws
    except Exception as e:
        print(f"WebSocket setup error: {e}")
        return None

# Send progress via global WebSocket
def send_progress(progress, message=''):
    global global_ws
    try:
        with ws_lock:
            if global_ws and global_ws.sock and global_ws.sock.connected:
                progress_data = {
                    'progress': round(progress, 2),
                    'message': message
                }
                global_ws.send(json.dumps(progress_data))
            else:
                print(f"Cannot send progress: WebSocket not connected. Progress: {progress}, Message: {message}")
    except Exception as e:
        print(f"Progress sending error: {e}")

# Coordinates for text placement on PDF
coordinates = {
    "Référence de certificat": (229.0, 650.0),
    "Date de création": (46.0, 600.0),
    "Numéro chez le déclarant": (180.0, 600.0),
    "YYYY de Date de paiement": (300.0, 600.0),
    "Date de paiement": (440.0, 600.0),
    "Declarant": (229.0, 545),
    "Id Aziza": (229.0, 533),
    "Nom et prenom ou raison sociale": (229.0, 510.0),
    "Adresse Aziza": (229.0, 495),
    "Preffession Aziza": (229.0, 475),
    "Identifiant du bénéficiaire": (229.0, 402.0),
    "Raison social du bénéficiaire": (229.0, 390.0),
    "type identifiant ben": (229.0, 420),
    "adresse benificiaire": (229.0, 360),
    "activite benificiaire": (229.0, 350),
    "totalMontantHT": (205.0, 260.0),
    "TVA_due": (260.0, 260.0),
    "Montant total TVA comprise - Total": (307.0, 245.0),
    "Montant total TVA comprise": (307.0, 260.0),
    "TVA retenue à la source": (360.0, 260.0),
    "Montant de la retenue": (461.0, 260.0),
    "Taux de la retenue": (408.0, 260.0),
    "Montant servie": (510.0, 260.0),
    "Montant servie - Total": (510.0, 245.0),
    "QR Code": (40.0, 120.0),
}

# Setup logging
def setup_logging():
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file = f"certificates_{current_datetime}.log"
    logging.basicConfig(filename=log_file, level=logging.INFO)
    logging.info('Started generating PDFs.')

# Function to reshape and display Arabic text correctly
def reshape_arabic(text):
    return get_display(arabic_reshaper.reshape(text))

# Register Arabic-supporting font
font_path = os.path.join('Amiri', 'Amiri-Regular.ttf')
pdfmetrics.registerFont(TTFont('Amiri', font_path))

def write_text_on_pdf(c, text, coordinates, font='Helvetica', font_size=4):
    if any('\u0600' <= char <= '\u06FF' for char in text):  # Check for Arabic characters
        bidi_text = reshape_arabic(text)  # Reshape and reorder Arabic text
        c.setFont('Amiri', font_size)
    else:
        c.setFont(font, font_size)
        bidi_text = text

    x, y = coordinates
    c.drawString(x, y, bidi_text)

# Function to generate PDF for a single row
def generate_pdf(row, progress_tracker, total_files, output_dir):
    try:
        os.makedirs(output_dir, exist_ok=True)

        if not any(row):
            logging.warning("Skipped empty row")
            return

        row = [str(cell).replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]

        # PDF generation logic
        reader = PdfReader("certificate.pdf")
        writer = PdfWriter()
        page = reader.pages[0]

        packet = BytesIO()
        can = canvas.Canvas(packet)

        # Write text fields from the Excel row
        write_text_on_pdf(can, str(row[3]), coordinates['Référence de certificat'], font_size=8)
        write_text_on_pdf(can, str(row[22]), coordinates['Date de création'], font_size=8)
        write_text_on_pdf(can, str(row[2]), coordinates['Numéro chez le déclarant'], font_size=8)
        write_text_on_pdf(can, str(row[4].split('-')[2]), coordinates['YYYY de Date de paiement'], font_size=8)
        write_text_on_pdf(can, str(row[4]), coordinates['Date de paiement'], font_size=8)

        # Information Déclarant
        write_text_on_pdf(can, 'Matricule fiscale', coordinates['Declarant'], font_size=8)
        write_text_on_pdf(can, '1259149J', coordinates['Id Aziza'], font_size=8)
        write_text_on_pdf(can, 'STE AZIZA DE COMMERCE DE DETAIL', coordinates['Nom et prenom ou raison sociale'], font_size=8)
        write_text_on_pdf(can, '022 ELECTRICITE Z IND BEN AROUS 2013', coordinates['Adresse Aziza'], font_size=8)
        write_text_on_pdf(can, 'VTE PDTS DIVERS', coordinates['Preffession Aziza'], font_size=8)

        # Information Béneficiaire
        write_text_on_pdf(can, str(row[19]), coordinates['type identifiant ben'], font_size=8)
        write_text_on_pdf(can, str(row[17]), coordinates['Identifiant du bénéficiaire'], font_size=8)
        write_text_on_pdf(can, str(row[18]), coordinates['Raison social du bénéficiaire'], font_size=8)
        write_text_on_pdf(can, str(row[23]), coordinates['adresse benificiaire'], font_size=8)
        write_text_on_pdf(can, str(row[24]), coordinates['activite benificiaire'], font_size=8)

        # tableau detail declaration
        write_text_on_pdf(can, str(row[7]), coordinates['totalMontantHT'], font_size=8)
        write_text_on_pdf(can, str(row[8]), coordinates['Montant total TVA comprise'], font_size=8)
        write_text_on_pdf(can, str(row[8]), coordinates['Montant total TVA comprise - Total'], font_size=8)
        write_text_on_pdf(can, '0', coordinates['TVA retenue à la source'], font_size=8)
        write_text_on_pdf(can, str(row[11]), coordinates['TVA_due'], font_size=8)
        write_text_on_pdf(can, str(row[8]), coordinates['Montant de la retenue'], font_size=8)
        write_text_on_pdf(can, str(row[10]), coordinates['Taux de la retenue'], font_size=8)
        write_text_on_pdf(can, str(row[12]), coordinates['Montant servie'], font_size=8)
        write_text_on_pdf(can, str(row[12]), coordinates['Montant servie - Total'], font_size=8)

        # QR Code text
        write_text_on_pdf(can, f"{row[0]}#{row[4].split('-')[2]}#{row[4].split('-')[1]}#1#{row[3]}", (60, 60), font_size=15)

        # Generate QR code
        qr_text = f"{row[0]}#{row[4].split('-')[2]}#{row[4].split('-')[1]}#1#{row[3]}"
        qr = qrcode.QRCode(box_size=10, border=2)
        qr.add_data(qr_text)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_byte_stream = BytesIO()
        qr_img.save(qr_byte_stream, format='PNG')
        qr_byte_stream.seek(0)
        qr_image = ImageReader(qr_byte_stream)

        can.drawImage(qr_image, 40, 120, width=100, height=100)
        can.save()

        packet.seek(0)
        overlay_pdf = PdfReader(packet)
        page.merge_page(overlay_pdf.pages[0])
        writer.add_page(page)

        output_pdf_name = os.path.join(output_dir, f"certificat_{row[17]}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf")
        with open(output_pdf_name, "wb") as output_pdf:
            writer.write(output_pdf)

        # Progress tracking
        processed_progress = progress_tracker.update_progress()
        if processed_progress is not None:
            send_progress(processed_progress, f"Generated PDF: {processed_progress:.2f}%")

    except Exception as e:
        logging.error(f"PDF Generation Error: {e}")
        processed_progress = progress_tracker.update_progress()
        if processed_progress is not None:
            send_progress(processed_progress, f"Error: {processed_progress:.2f}%")

def process_files(excel_file, socket_id):
    setup_websocket(socket_id)
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active
    
    total_files = sheet.max_row - 1
    logging.basicConfig(level=logging.INFO)
    
    progress_tracker = ProgressTracker(total_files)
    
    # Create datetime-based output directory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f'output_{timestamp}'
    os.makedirs(output_dir, exist_ok=True)

    # Create outputs directory
    output_folder = os.path.join(os.getcwd(), "outputs")
    os.makedirs(output_folder, exist_ok=True)

    send_progress(0, "Starting PDF generation")

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
        futures = [
            executor.submit(generate_pdf, row, progress_tracker, total_files, output_dir)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)
        ]
        
        for future in as_completed(futures):
            future.result()

    send_progress(95, "Creating ZIP archive")

    # Store ZIP file in outputs/ folder
    zip_filename = os.path.join(output_folder, f'output_{timestamp}.zip')
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, output_dir)
                zipf.write(file_path, arcname)

    # Clean up the original directory after zipping
    shutil.rmtree(output_dir)

    send_progress(100, "PDF generation and ZIP creation completed")

    total_time = time.time() - start_time
    print(f"\nTotal processing time: {total_time / 60:.2f} minutes")
    print(f"Output file: {zip_filename}")

    return zip_filename


if __name__ == "__main__":
    excel_file = sys.argv[1]
    socket_id = sys.argv[2]
    process_files(excel_file, socket_id)